#!/usr/bin/env python3
"""
Build the merchant-facing product data workbook.

    python3 tools/build-product-workbook.py

Inputs
    product-data/product-data.json        56 products, captured 2026-07-27
    data-schema/reference-fragrances.json the 40 originals, researched 2026-08-13
Spec
    data-schema/*.md
Output
    data-schema/siwa-product-data.xlsx

Design rule: the owner is only ever asked for what cannot be worked out.
Anything derivable from another answer is computed on the "Auto-calculated"
sheet with a live formula, so it is visible rather than hidden.

Vocabularies are Fragrantica's: scent families are Fragrantica main-accord
names, and the note list is built from the actual pyramids of the 40
fragrances this catalogue references.
"""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "product-data" / "product-data.json"
REF = ROOT / "data-schema" / "reference-fragrances.json"
OUT = ROOT / "data-schema" / "siwa-product-data.xlsx"

# ---------------------------------------------------------------- style
HEAD_BG, HEAD_FG = "212012", "FFFFFF"
GIVEN_BG = "EFECE4"     # already known — locked
INPUT_BG = "FFFFFF"     # you type here
DRAFT_BG = "FFF8E7"     # pre-filled suggestion — check it
AUTO_BG = "EAF1EA"      # calculated
CANVAS = "F7F5EE"
MUTED = "7D6A5B"
ACCENT = "B18044"

THIN = Side(style="thin", color="CCBCA0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------- vocabularies
# Fragrantica main accords. ✔ = seen directly in the 2026-08-13 research of the
# 40 referenced fragrances (data-schema/reference-fragrances.json); the rest are
# standard Fragrantica accords carried for coverage.
FAMILIES = [
    ("Amber", "✔"), ("Animalic", ""), ("Anis", ""), ("Aquatic", "✔"),
    ("Aromatic", "✔"), ("Balsamic", ""), ("Cacao", ""), ("Caramel", ""),
    ("Cherry", ""), ("Chypre", "✔"), ("Cinnamon", ""), ("Citrus", "✔"),
    ("Coconut", "✔"), ("Conifer", ""), ("Earthy", ""), ("Floral", "✔"),
    ("Fresh", "✔"), ("Fresh Spicy", ""), ("Fruity", "✔"), ("Gourmand", "✔"),
    ("Green", ""), ("Herbal", ""), ("Honey", ""), ("Iris", ""),
    ("Lactonic", ""), ("Lavender", ""), ("Leather", ""), ("Marine", ""),
    ("Mossy", ""), ("Musky", "✔"), ("Nutty", "✔"), ("Oud", "✔"),
    ("Ozonic", ""), ("Patchouli", ""), ("Powdery", "✔"), ("Rose", "✔"),
    ("Salty", "✔"), ("Smoky", ""), ("Soft Spicy", ""), ("Sweet", "✔"),
    ("Tobacco", ""), ("Tropical", ""), ("Vanilla", "✔"), ("Violet", ""),
    ("Warm Spicy", "✔"), ("White Floral", "✔"), ("Woody", "✔"), ("Yellow Floral", ""),
]

INTENSITY = [
    ("Subtle", "stays close to the skin"),
    ("Distinct", "noticeable without dominating"),
    ("Powerful", "fills the room"),
]
SILLAGE = [
    ("Subtle", "trail of about an arm's length"),
    ("Distinct", "trail across a table"),
    ("Powerful", "trail across a room"),
]
GENDER = [("Masculine", ""), ("Feminine", ""), ("Unisex", "")]
CONCENTRATION = [
    ("Extrait de Parfum", "what the bottles are printed with"),
    ("Eau de Parfum", "what the website currently says"),
]
YESNO = [("Yes", ""), ("No", "")]

# family -> season. Applied by the Auto-calculated sheet; override by typing over it.
SEASON_MAP = {
    "Aquatic": "Summer", "Marine": "Summer", "Citrus": "Summer", "Fresh": "Summer",
    "Salty": "Summer", "Ozonic": "Summer", "Tropical": "Summer", "Coconut": "Summer",
    "Floral": "Spring", "White Floral": "Spring", "Yellow Floral": "Spring",
    "Green": "Spring", "Rose": "Spring", "Fruity": "Spring", "Iris": "Spring",
    "Violet": "Spring", "Herbal": "Spring", "Lavender": "Spring", "Aromatic": "Spring",
    "Woody": "Autumn", "Chypre": "Autumn", "Mossy": "Autumn", "Earthy": "Autumn",
    "Leather": "Autumn", "Powdery": "Autumn", "Musky": "Autumn", "Fresh Spicy": "Autumn",
    "Patchouli": "Autumn", "Conifer": "Autumn", "Anis": "Autumn", "Lactonic": "Autumn",
    "Amber": "Winter", "Gourmand": "Winter", "Vanilla": "Winter", "Sweet": "Winter",
    "Oud": "Winter", "Tobacco": "Winter", "Warm Spicy": "Winter", "Smoky": "Winter",
    "Balsamic": "Winter", "Caramel": "Winter", "Cacao": "Winter", "Honey": "Winter",
    "Nutty": "Winter", "Cherry": "Winter", "Cinnamon": "Winter", "Animalic": "Winter",
    "Soft Spicy": "Winter",
}

# ---------------------------------------------------------------- product line
BUNDLES = {"layering-30-ml-bundle", "marshmallow-bundle", "vanilla-bundle"}
BODY_CARE = {"apple-pie", "marshmallow", "silk-vanilla", "sweet-rum", "vanilla-91",
             "silk-vanilla-body-lotion"}
LAYERING = {"layering-apple", "layering-lychee", "layering-pistachio", "layering-vanilla"}
LINE_LABEL = {"originals": "Originals", "inspired-by": "Inspired by",
              "layering": "Layering", "body-care": "Body care", "bundle": "Bundle"}


def product_line(p):
    h = p["handle"]
    if h in BUNDLES:
        return "bundle"
    if h in BODY_CARE:
        return "body-care"
    if h in LAYERING:
        return "layering"
    return "originals" if p["vendor_kind"] == "siwa_owned" else "inspired-by"


def is_original(line):
    return line in ("originals", "body-care", "bundle")


def gender_from_tags(p):
    """Explicit Unisex wins. Men+Women together is ambiguous — leave for a human."""
    t = set(p["tags"])
    if "Unisex" in t:
        return "Unisex"
    if "Men" in t and "Women" in t:
        return ""
    if "Men" in t:
        return "Masculine"
    if "Women" in t:
        return "Feminine"
    return ""


TIER_RE = re.compile(
    r"(top|heart|middle|base)\s*notes?\s*:?\s*(.+?)(?=(?:top|heart|middle|base)\s*notes?\s*:|best for|$)",
    re.I | re.S)


def notes_from_store(p):
    """Whatever tiers the merchant already wrote, normalised across the 13 spellings."""
    out = {"top": "", "heart": "", "base": ""}
    for tier, text in (p.get("notes") or {}).items():
        key = "heart" if tier in ("heart", "middle") else tier
        if key in out and text:
            out[key] = text.strip()
    return out


# ---------------------------------------------------------------- helpers
def header(ws, labels, widths, height=40, freeze="C2"):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = Font(bold=True, color=HEAD_FG, size=10)
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = height
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze


def paint(ws, first, last, spec):
    """spec: {column_index: 'given' | 'input' | 'draft' | 'auto'}"""
    fills = {"given": GIVEN_BG, "input": INPUT_BG, "draft": DRAFT_BG, "auto": AUTO_BG}
    for r in range(first, last + 1):
        for cidx, kind in spec.items():
            c = ws.cell(row=r, column=cidx)
            c.border = BORDER
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.fill = PatternFill("solid", fgColor=fills[kind])
            c.protection = c.protection.copy(locked=kind in ("given", "auto"))


def dropdown(ws, name, rng, strict=True, prompt=None):
    dv = DataValidation(type="list", formula1=f"={name}", allow_blank=True,
                        showErrorMessage=strict)
    if strict:
        dv.errorTitle, dv.error = "Not on the list", "Pick a value from the dropdown."
    if prompt:
        dv.promptTitle, dv.prompt, dv.showInputMessage = "How to fill this", prompt, True
    ws.add_data_validation(dv)
    dv.add(rng)


def note(ws, row, col, text, span=1, color=MUTED):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(size=10, color=color, italic=True)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)


def main():
    products = json.loads(SRC.read_text(encoding="utf-8"))
    refs = json.loads(REF.read_text(encoding="utf-8"))
    refs.pop("_meta", None)
    products.sort(key=lambda p: (not is_original(product_line(p)), p["handle"]))
    n = len(products)

    wb = Workbook()
    wb.remove(wb.active)          # drop the default empty sheet
    defined = {}

    # ============================================================ Lists
    ws_l = wb.create_sheet("Lists")
    col = 1

    def put(title, values, name, note_col=True, width=(24, 34)):
        nonlocal col
        h = ws_l.cell(row=1, column=col, value=title)
        h.font = Font(bold=True, color=HEAD_FG, size=10)
        h.fill = PatternFill("solid", fgColor=HEAD_BG)
        h.alignment = Alignment(horizontal="center", wrap_text=True)
        ws_l.column_dimensions[get_column_letter(col)].width = width[0]
        if note_col:
            h2 = ws_l.cell(row=1, column=col + 1, value="—")
            h2.font = Font(bold=True, color=HEAD_FG, size=10)
            h2.fill = PatternFill("solid", fgColor=HEAD_BG)
            ws_l.column_dimensions[get_column_letter(col + 1)].width = width[1]
        for i, item in enumerate(values, start=2):
            val, meaning = item if isinstance(item, tuple) else (item, "")
            ws_l.cell(row=i, column=col, value=val).font = Font(size=10)
            if note_col:
                ws_l.cell(row=i, column=col + 1, value=meaning).font = Font(size=9, color=MUTED)
        defined[name] = (f"'Lists'!${get_column_letter(col)}$2:"
                         f"${get_column_letter(col)}${len(values) + 1}")
        col += 2 if note_col else 1

    put("Scent family (Fragrantica accord)", FAMILIES, "LIST_FAMILY", width=(28, 8))
    put("Intensity", INTENSITY, "LIST_INTENSITY", width=(16, 32))
    put("Sillage", SILLAGE, "LIST_SILLAGE", width=(16, 32))
    put("Gender", GENDER, "LIST_GENDER", width=(16, 6))
    put("Concentration", CONCENTRATION, "LIST_CONC", width=(22, 34))
    put("Yes / No", YESNO, "LIST_YESNO", width=(10, 6))

    # Note vocabulary, built from the pyramids of the 40 referenced fragrances
    vocab = set()
    for r in refs.values():
        for tier in ("top", "heart", "base"):
            vocab.update(r.get(tier) or [])
    put("Fragrance notes", sorted(vocab), "LIST_NOTES", note_col=False, width=(30, 0))

    season_pairs = sorted(SEASON_MAP.items())
    put("Family", [k for k, _ in season_pairs], "LIST_SEASONKEY", note_col=False, width=(20, 0))
    put("→ Season", [v for _, v in season_pairs], "LIST_SEASONVAL", note_col=False, width=(14, 0))
    season_first = get_column_letter(col - 2)
    season_last = get_column_letter(col - 1)
    defined["MAP_SEASON"] = (f"'Lists'!${season_first}$2:${season_last}${len(season_pairs) + 1}")

    for name, ref in defined.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))
    ws_l.sheet_view.showGridLines = False
    note(ws_l, len(vocab) + 4, 1,
         "Scent families are Fragrantica main-accord names; ✔ marks the ones observed directly "
         "in the 2026-08-13 research of the 40 referenced fragrances. The note list is every note "
         "appearing in those 40 pyramids — it is a starting vocabulary, and the note dropdowns "
         "accept new values.", span=4)
    ws_l.protection.sheet = True

    # ============================================================ Read me
    ws0 = wb.create_sheet("Read me", 0)
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 3
    for cl in "BCDEFG":
        ws0.column_dimensions[cl].width = 19

    t = ws0.cell(row=2, column=2, value="Siwa Fragrances — product data")
    t.font = Font(size=22, bold=True, color=HEAD_BG)
    ws0.merge_cells("B2:G2")
    s = ws0.cell(row=3, column=2,
                 value="Three sheets to fill. Everything else is worked out from them.")
    s.font = Font(size=12, color=MUTED)
    ws0.merge_cells("B3:G3")

    blocks = [
        ("What to fill, in order", [
            "Settings ......... 3 answers that apply to the whole catalogue. Do these first.",
            "Products ......... one row per product. The only long sheet.",
            "Heritage ......... 16 rows. Originals only — the story behind each one.",
            "",
            "Auto-calculated .. nothing to do. Shows what the three sheets above produce.",
            "Progress ......... nothing to do. Tracks how far along you are.",
            "Lists ............ the source of every dropdown. Leave it alone.",
        ]),
        ("Cell colours", [
            "Grey    Already known from your store. Locked.",
            "Cream   Filled in for you as a suggestion. Read it, correct it if it is wrong.",
            "White   Yours to write.",
            "Green   Calculated. Do not retype it unless you want to override it.",
        ]),
        ("Where the cream suggestions came from", [
            "Every product on the Inspired by line references a real designer fragrance, and that",
            "fragrance's composition is public. Those pyramids were looked up on 13 Aug 2026 and",
            "filled in for you. They describe the ORIGINAL, not your version — so read each one",
            "and correct it to match what you actually blended.",
            "",
            "One check on the method: Sundaze is the only product whose own description already",
            "lists its notes. The researched pyramid for Armani Power Of You matched it exactly.",
        ]),
        ("The one rule that cannot bend", [
            "Heritage — the oasis, the Siwan craft, the Arabic naming — belongs only to the 16",
            "products Siwa makes itself. The 40 inspired-by products stay commercial.",
            "That is why Heritage is a separate 16-row sheet and not a set of columns you could",
            "fill in by accident.",
            "",
            "Note that the four Layering products count as inspired-by: they reference Kayali.",
        ]),
        ("What is NOT asked, because it is worked out", [
            "Register, concentration per product, season, product type, badge, launch date,",
            "image alt text, card descriptors, tags, cross-sells, JSON-LD brand.",
            "All of it is on the Auto-calculated sheet, with the formula visible.",
        ]),
    ]

    r = 5
    for title, lines in blocks:
        h = ws0.cell(row=r, column=2, value=title)
        h.font = Font(size=13, bold=True, color=HEAD_BG)
        h.fill = PatternFill("solid", fgColor=CANVAS)
        ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws0.row_dimensions[r].height = 22
        r += 1
        for line in lines:
            c = ws0.cell(row=r, column=2, value=line)
            c.font = Font(size=10, name="Consolas" if "." * 3 in line else None)
            c.alignment = Alignment(vertical="center")
            ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            r += 1
        r += 1

    for i, (label, colr) in enumerate(
            [("Grey", GIVEN_BG), ("Cream", DRAFT_BG), ("White", INPUT_BG), ("Green", AUTO_BG)]):
        cc = ws0.cell(row=13 + i, column=8, value=label)
        cc.fill = PatternFill("solid", fgColor=colr)
        cc.border = BORDER
        cc.font = Font(size=9)
    ws0.column_dimensions["H"].width = 12
    ws0.protection.sheet = True

    # ============================================================ Settings
    ws_s = wb.create_sheet("Settings", 1)
    ws_s.sheet_view.showGridLines = False
    ws_s.column_dimensions["A"].width = 3
    ws_s.column_dimensions["B"].width = 46
    ws_s.column_dimensions["C"].width = 26
    ws_s.column_dimensions["D"].width = 62

    t = ws_s.cell(row=2, column=2, value="Catalogue-wide answers")
    t.font = Font(size=16, bold=True, color=HEAD_BG)
    ws_s.cell(row=3, column=2,
              value="Asked once here instead of 56 times on the Products sheet."
              ).font = Font(size=10, color=MUTED)

    rows = [
        ("Concentration for the perfumes", "",
         "Your bottles say EXTRAIT DE PARFUM. Your website says Eau de Parfum. "
         "Pick the true one — it drives the product type, the spec line and Google.",
         "LIST_CONC"),
        ("Reviews needed for a Best seller badge", 20,
         "A product with at least this many reviews gets the badge automatically.", None),
        ("Show the original's retail price?", "No",
         "Leave as No. Naming a designer price alongside yours is a legal question, not a "
         "design one, and it has not been answered yet.", "LIST_YESNO"),
    ]
    for i, (label, default, why, listname) in enumerate(rows):
        r = 5 + i * 2
        lc = ws_s.cell(row=r, column=2, value=label)
        lc.font = Font(size=11, bold=True)
        lc.alignment = Alignment(vertical="center")
        vc = ws_s.cell(row=r, column=3, value=default)
        vc.fill = PatternFill("solid", fgColor=INPUT_BG)
        vc.border = BORDER
        vc.font = Font(size=11)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.protection = vc.protection.copy(locked=False)
        wc = ws_s.cell(row=r, column=4, value=why)
        wc.font = Font(size=9, color=MUTED)
        wc.alignment = Alignment(vertical="center", wrap_text=True)
        ws_s.row_dimensions[r].height = 34
        if listname:
            dropdown(ws_s, listname, f"C{r}:C{r}")

    wb.defined_names.add(DefinedName("SET_CONC", attr_text="'Settings'!$C$5"))
    wb.defined_names.add(DefinedName("SET_BESTSELLER", attr_text="'Settings'!$C$7"))
    wb.defined_names.add(DefinedName("SET_SHOWPRICE", attr_text="'Settings'!$C$9"))
    ws_s.cell(row=12, column=2, value="Brand name published to Google").font = Font(size=11, bold=True)
    bc = ws_s.cell(row=12, column=3, value="Siwa Fragrances")
    bc.fill = PatternFill("solid", fgColor=GIVEN_BG)
    bc.border = BORDER
    bc.alignment = Alignment(horizontal="center")
    ws_s.cell(row=12, column=4,
              value="Fixed. Today Google is told the brand of 40 of your products is the designer "
                    "house named in the vendor field — that is the single most damaging data "
                    "problem on the store.").font = Font(size=9, color=MUTED)
    ws_s.row_dimensions[12].height = 34
    ws_s.protection.sheet = True
    ws_s.protection.selectLockedCells = False

    # ============================================================ Products
    ws = wb.create_sheet("Products", 2)
    cols = ["handle", "Product", "Line", "References", "The original's composition — work from this",
            "Top notes", "Heart notes", "Base notes",
            "Family 1", "Family 2", "Family 3",
            "Intensity", "Sillage", "Lasts (hours)", "Gender",
            "Arabic name", "Description (English)", "Description (Arabic)"]
    header(ws, cols,
           [16, 20, 12, 30, 46, 28, 28, 28, 15, 15, 15, 13, 13, 12, 13, 16, 40, 40], height=46)

    for i, p in enumerate(products):
        r = i + 2
        line = product_line(p)
        ref = refs.get(p["handle"])
        store = notes_from_store(p)

        ws.cell(row=r, column=1, value=p["handle"])
        ws.cell(row=r, column=2, value=p["title"])
        ws.cell(row=r, column=3, value=LINE_LABEL[line])

        context = []
        if ref:
            ws.cell(row=r, column=4, value=f"{ref['house']} — {ref['fragrance']}")
            comp = " · ".join(f"{t.capitalize()}: {', '.join(ref[t])}"
                              for t in ("top", "heart", "base") if ref.get(t))
            context.append(f"THE ORIGINAL — {comp}" if comp else "THE ORIGINAL — pyramid not published")
            if ref.get("note"):
                context.append(f"⚠ {ref['note']}")
            # Structured note names beat the store's prose sentences, so research wins here
            # and the merchant's own wording is preserved in the context column instead.
            for cidx, tier in ((6, "top"), (7, "heart"), (8, "base")):
                vals = ", ".join(ref.get(tier) or []) or store[tier]
                ws.cell(row=r, column=cidx, value=vals or None)
            for k, acc in enumerate(ref.get("accords", [])[:3]):
                ws.cell(row=r, column=9 + k, value=acc)
        else:
            ws.cell(row=r, column=4, value="Siwa original — nothing to reference")
            for cidx, tier in ((6, "top"), (7, "heart"), (8, "base")):
                if store[tier]:
                    ws.cell(row=r, column=cidx, value=store[tier])

        if any(store.values()):
            context.append("YOUR TEXT — " + " · ".join(
                f"{t.capitalize()}: {v}" for t, v in store.items() if v))
        elif p.get("body_text"):
            context.append("YOUR DESCRIPTION — " + p["body_text"][:220])
        ws.cell(row=r, column=5, value="\n".join(context) or None)

        ws.cell(row=r, column=15, value=gender_from_tags(p))
        ws.row_dimensions[r].height = 30

    paint(ws, 2, n + 1, {1: "given", 2: "given", 3: "given", 4: "given", 5: "given",
                         6: "draft", 7: "draft", 8: "draft",
                         9: "draft", 10: "draft", 11: "draft",
                         12: "input", 13: "input", 14: "input",
                         15: "draft", 16: "input", 17: "input", 18: "input"})

    for rng in (f"F2:H{n+1}",):
        dropdown(ws, "LIST_NOTES", rng, strict=False,
                 prompt="Comma-separated. Pick from the list or type your own.")
    dropdown(ws, "LIST_FAMILY", f"I2:K{n+1}",
             prompt="Fragrantica main accords. Family 1 is the dominant one — it also sets the season.")
    dropdown(ws, "LIST_INTENSITY", f"L2:L{n+1}")
    dropdown(ws, "LIST_SILLAGE", f"M2:M{n+1}")
    dropdown(ws, "LIST_GENDER", f"O2:O{n+1}")
    dv_h = DataValidation(type="whole", operator="between", formula1=1, formula2=24,
                          allow_blank=True, showErrorMessage=True)
    dv_h.errorTitle, dv_h.error = "Hours", "A whole number of hours, 1 to 24."
    ws.add_data_validation(dv_h)
    dv_h.add(f"N2:N{n+1}")

    for rng in (f"L2:L{n+1}", f"M2:M{n+1}", f"N2:N{n+1}", f"I2:I{n+1}"):
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'ISBLANK({rng.split(":")[0]})'],
            fill=PatternFill("solid", bgColor="FDF3F2"), stopIfTrue=False))

    ws.protection.sheet = True
    ws.protection.selectLockedCells = False

    # ============================================================ Heritage
    originals = [p for p in products if is_original(product_line(p))]
    m = len(originals)
    ws_h = wb.create_sheet("Heritage", 3)
    hcols = ["handle", "Product", "Where it comes from (EN)", "Where it comes from (AR)",
             "Ingredient origin (EN)", "Ingredient origin (AR)",
             "Who made it (EN)", "Who made it (AR)",
             "What the name means (EN)", "What the name means (AR)"]
    header(ws_h, hcols, [16, 20] + [34] * 8, height=44)

    for i, p in enumerate(originals):
        r = i + 2
        ws_h.cell(row=r, column=1, value=p["handle"])
        ws_h.cell(row=r, column=2, value=p["title"])
        ws_h.row_dimensions[r].height = 30
    paint(ws_h, 2, m + 1, {1: "given", 2: "given", **{c: "input" for c in range(3, 11)}})

    note(ws_h, m + 3, 1,
         "Only the 16 products Siwa makes itself are on this sheet. The 40 inspired-by products "
         "are absent by design — heritage language on a designer clone reads as decoration and "
         "the corpus rules it out.", span=10)
    note(ws_h, m + 5, 1,
         "\"Who made it\" is for a real collaboration only. If no Siwan artisan actually worked "
         "on the product, leave it empty — an empty cell is honest, a claim is not.", span=10, color="8F3A2E")
    ws_h.protection.sheet = True
    ws_h.protection.selectLockedCells = False

    # ============================================================ Auto-calculated
    ws_a = wb.create_sheet("Auto-calculated", 4)
    acols = ["handle", "Register", "Concentration", "Product type", "Season",
             "Badge", "Launch date", "Card descriptors", "Image alt text", "Tags"]
    header(ws_a, acols, [16, 13, 20, 18, 12, 14, 13, 26, 46, 40], height=40)

    cat = {}
    for p in products:
        h, line = p["handle"], product_line(p)
        if line == "bundle":
            cat[h] = "Bundle"
        elif h == "silk-vanilla-body-lotion":
            cat[h] = "Body Lotion"
        elif line == "body-care":
            cat[h] = "Body Splash"
        else:
            cat[h] = None

    for i, p in enumerate(products):
        r = i + 2
        pr = r  # same row on Products, both sheets share the sort order
        fixed = cat[p["handle"]]
        ws_a.cell(row=r, column=1, value=p["handle"])
        ws_a.cell(row=r, column=2,
                  value=f'=IF(OR(Products!C{pr}="Inspired by",Products!C{pr}="Layering"),'
                        f'"Inspired","Original")')
        ws_a.cell(row=r, column=3,
                  value=f'="{fixed}"' if fixed else "=SET_CONC")
        ws_a.cell(row=r, column=4, value=f"=C{r}")
        ws_a.cell(row=r, column=5,
                  value=f'=IFERROR(VLOOKUP(Products!I{pr},MAP_SEASON,2,FALSE),"All season")')
        ws_a.cell(row=r, column=6,
                  value=f'=IF({p["review_count"] or 0}>=SET_BESTSELLER,"best-seller","")')
        ws_a.cell(row=r, column=7, value=p["published_at"])
        ws_a.cell(row=r, column=8,
                  value=f'=TEXTJOIN(", ",TRUE,Products!I{pr},Products!J{pr},Products!K{pr})')
        ws_a.cell(row=r, column=9,
                  value=f'=TEXTJOIN(" — ",TRUE,Products!B{pr},Products!P{pr},C{r})&'
                        f'" — Siwa Fragrances"')
        ws_a.cell(row=r, column=10,
                  value=f'=TEXTJOIN(", ",TRUE,"line:"&LOWER(SUBSTITUTE(Products!C{pr}," ","-")),'
                        f'IF(F{r}="","","status:"&F{r}),'
                        f'IF(Products!O{pr}="","","gender:"&LOWER(Products!O{pr})),'
                        f'IF(E{r}="","","season:"&LOWER(E{r})))')
    paint(ws_a, 2, n + 1, {c: "auto" for c in range(1, 11)})
    note(ws_a, n + 3, 1,
         "Nothing on this sheet is typed. Each column is a formula over Products, Heritage and "
         "Settings — shown rather than hidden, so you can see that dropping those questions did "
         "not drop the data. Season comes from Family 1 through the mapping table on Lists; type "
         "over a cell to override it.", span=10)
    ws_a.protection.sheet = True

    # ============================================================ Progress
    ws_p = wb.create_sheet("Progress", 5)
    header(ws_p, ["Sheet", "Field", "Needed", "Done", "%"], [18, 34, 12, 12, 12],
           height=26, freeze="A2")
    # Originals sort first on Products, so rows 2..m+1 there are exactly the 16 originals.
    tracked = [
        ("Products", "Top notes", "F", n, n + 1),
        ("Products", "Heart notes", "G", n, n + 1),
        ("Products", "Base notes", "H", n, n + 1),
        ("Products", "Family 1", "I", n, n + 1),
        ("Products", "Intensity", "L", n, n + 1),
        ("Products", "Sillage", "M", n, n + 1),
        ("Products", "Lasts (hours)", "N", n, n + 1),
        ("Products", "Gender", "O", n, n + 1),
        ("Products", "Arabic name (originals)", "P", m, m + 1),
        ("Products", "Description (English)", "Q", n, n + 1),
        ("Products", "Description (Arabic)", "R", n, n + 1),
        ("Heritage", "Where it comes from (EN)", "C", m, m + 1),
        ("Heritage", "What the name means (EN)", "I", m, m + 1),
    ]
    for i, (sheet, label, cl, total, rows_end) in enumerate(tracked):
        r = i + 2
        ws_p.cell(row=r, column=1, value=sheet)
        ws_p.cell(row=r, column=2, value=label)
        ws_p.cell(row=r, column=3, value=total)
        ws_p.cell(row=r, column=4, value=f"=COUNTA('{sheet}'!{cl}2:{cl}{rows_end})")
        pc = ws_p.cell(row=r, column=5, value=f"=IF(C{r}=0,0,MIN(1,D{r}/C{r}))")
        pc.number_format = "0%"
        for cidx in range(1, 6):
            c = ws_p.cell(row=r, column=cidx)
            c.border, c.font = BORDER, Font(size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
    last = len(tracked) + 1
    ws_p.conditional_formatting.add(f"E2:E{last}", CellIsRule(
        operator="greaterThanOrEqual", formula=["1"], fill=PatternFill("solid", bgColor="D9EAD3")))
    ws_p.conditional_formatting.add(f"E2:E{last}", CellIsRule(
        operator="lessThan", formula=["0.5"], fill=PatternFill("solid", bgColor="FDF3F2")))
    note(ws_p, last + 2, 1,
         "The note rows start above zero because the inspired-by pyramids were researched and "
         "filled in for you. Check them before counting them as done.", span=5)
    ws_p.protection.sheet = True

    wb.move_sheet("Lists", offset=len(wb.sheetnames))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    filled = sum(1 for p in products if refs.get(p["handle"]))
    print(f"✅ {OUT}")
    print(f"   {n} products · {m} originals · {filled} pyramids pre-filled from research")
    print(f"   {len(vocab)} notes · {len(FAMILIES)} Fragrantica accords · "
          f"{len(wb.sheetnames)} sheets")
    print(f"   sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
